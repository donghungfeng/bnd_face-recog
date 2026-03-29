from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date, timedelta
import openpyxl
import io
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import joinedload
from sqlalchemy.orm import aliased

from fastapi.templating import Jinja2Templates
templates = Jinja2Templates(directory="templates")

from database import get_db
from models import ShiftCategory, ShiftAssignment, Employee, OrganizationUnit
from schemas import ShiftCategoryCreate, ShiftAssignmentCreate
from routers.auth_router import get_current_user

router = APIRouter()

@router.get("/assignments")
def read_assignments(request: Request): 
    return templates.TemplateResponse("assignments.html", {"request": request})

# ==========================================
# 1. API DANH MỤC CA TRỰC
# ==========================================
@router.post("/api/shifts")
def create_shift_category(shift: ShiftCategoryCreate, db: Session = Depends(get_db)):
    db_shift = db.query(ShiftCategory).filter(ShiftCategory.shift_code == shift.shift_code).first()
    if db_shift:
        raise HTTPException(status_code=400, detail="Mã ca trực đã tồn tại")
    
    new_shift = ShiftCategory(**shift.dict())
    db.add(new_shift)
    db.commit()
    return {"status": "success", "message": "Thêm ca trực thành công"}

@router.get("/api/shifts")
def get_shifts(db: Session = Depends(get_db)):
    shifts = db.query(ShiftCategory).all()
    result = []
    for s in shifts:
        result.append({
            "id": s.id,
            "shift_code": s.shift_code,
            "shift_name": s.shift_name,
            "start_time": s.start_time.strftime("%H:%M") if s.start_time else None,
            "end_time": s.end_time.strftime("%H:%M") if s.end_time else None,
            "is_overnight": s.is_overnight,
            "status": s.status,
            "notes": s.notes,
            "checkin_from": s.checkin_from.strftime("%H:%M") if s.checkin_from else None,
            "checkin_to": s.checkin_to.strftime("%H:%M") if s.checkin_to else None,
            "checkout_from": s.checkout_from.strftime("%H:%M") if s.checkout_from else None,
            "checkout_to": s.checkout_to.strftime("%H:%M") if s.checkout_to else None,
            "work_hours": s.work_hours,
            "work_days": s.work_days,
            "day_coefficient": s.day_coefficient
        })
    return result

@router.put("/api/shifts/{shift_code}")
def update_shift_category(shift_code: str, shift: ShiftCategoryCreate, db: Session = Depends(get_db)):
    db_shift = db.query(ShiftCategory).filter(ShiftCategory.shift_code == shift_code).first()
    if not db_shift:
        raise HTTPException(status_code=404, detail="Không tìm thấy ca trực")
    
    db_shift.shift_name = shift.shift_name
    db_shift.start_time = shift.start_time
    db_shift.end_time = shift.end_time
    db_shift.is_overnight = shift.is_overnight
    db_shift.status = shift.status
    db_shift.notes = shift.notes
    db_shift.checkin_from = shift.checkin_from
    db_shift.checkin_to = shift.checkin_to
    db_shift.checkout_from = shift.checkout_from
    db_shift.checkout_to = shift.checkout_to
    db_shift.work_hours = shift.work_hours
    db_shift.work_days = shift.work_days
    db_shift.day_coefficient = shift.day_coefficient

    db.commit()
    return {"status": "success", "message": "Cập nhật thành công"}

@router.get("/api/shifts/{shift_code}")
def delete_shift(shift_code: str, db: Session = Depends(get_db)):
    shift = db.query(ShiftCategory).filter(ShiftCategory.shift_code == shift_code).first()
    if not shift: 
        raise HTTPException(status_code=404, detail="Không tìm thấy ca trực")
    db.delete(shift)
    db.commit()
    return {"status": "success"}


# ==========================================
# 2. API PHÂN CÔNG CA TRỰC & EXCEL
# ==========================================
@router.get("/api/assignments")
def get_assignments(month: int, year: int, db: Session = Depends(get_db)):
    start_date = date(year, month, 1)
    if month == 12: 
        end_date = date(year + 1, 1, 1)
    else: 
        end_date = date(year, month + 1, 1)

    assignments = db.query(ShiftAssignment, Employee.username, Employee.full_name, ShiftCategory.shift_name).join(
        Employee, ShiftAssignment.employee_id == Employee.id
    ).join(
        ShiftCategory, ShiftAssignment.shift_code == ShiftCategory.shift_code
    ).filter(
        ShiftAssignment.shift_date >= start_date,
        ShiftAssignment.shift_date < end_date
    ).all()

    return [{
        "id": a[0].id, 
        "employee_id": a[0].employee_id,
        "username": a[1], 
        "full_name": a[2],
        "shift_code": a[0].shift_code, 
        "shift_name": a[3], 
        "shift_date": a[0].shift_date,
        "assigner": a[0].assigner
    } for a in assignments]

@router.post("/api/assignments")
def create_assignment(req: ShiftAssignmentCreate, db: Session = Depends(get_db)):
    existing = db.query(ShiftAssignment).filter(
        ShiftAssignment.employee_id == req.employee_id, 
        ShiftAssignment.shift_date == req.shift_date
    ).first()
    
    if existing:
        existing.shift_code = req.shift_code 
        if req.assigner: existing.assigner = req.assigner
    else:
        db.add(ShiftAssignment(**req.dict()))
    db.commit()
    return {"status": "success"}

@router.delete("/api/assignments/{assign_id}")
def delete_assignment(assign_id: int, db: Session = Depends(get_db)):
    item = db.query(ShiftAssignment).filter(ShiftAssignment.id == assign_id).first()
    if item:
        db.delete(item)
        db.commit()
    return {"status": "success"}

@router.get("/api/assignments/details")
def get_assignments_details(start_date: str = None, end_date: str = None, month: int = None, year: int = None, db: Session = Depends(get_db)):
    from sqlalchemy.orm import aliased
    Dept = aliased(OrganizationUnit)
    ParentUnit = aliased(OrganizationUnit)

    query = db.query(
        ShiftAssignment.id,
        ShiftAssignment.shift_code,
        ShiftAssignment.shift_date,
        Employee.id.label("employee_id"),
        Employee.full_name,
        Employee.dob,
        Employee.date_of_birth,
        Dept.unit_name.label("ten_phong_ban"),
        ParentUnit.unit_name.label("ten_don_vi")
    ).join(
        Employee, ShiftAssignment.employee_id == Employee.id
    ).outerjoin(
        Dept, Employee.department_id == Dept.id
    ).outerjoin(
        ParentUnit, Dept.parent_id == ParentUnit.id
    )

    if start_date and end_date:
        query = query.filter(
            ShiftAssignment.shift_date >= datetime.strptime(start_date, "%Y-%m-%d").date(),
            ShiftAssignment.shift_date <= datetime.strptime(end_date, "%Y-%m-%d").date()
        )
    elif month and year:
        s_date = date(year, month, 1)
        e_date = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
        query = query.filter(
            ShiftAssignment.shift_date >= s_date,
            ShiftAssignment.shift_date < e_date
        )

    results = query.all()

    data = []
    for row in results:
        dob_val = row.dob or row.date_of_birth
        ten_don_vi = row.ten_don_vi or row.ten_phong_ban or ""
        ten_phong_ban = row.ten_phong_ban or ""

        data.append({
            "id": row.id,
            "employee_id": row.employee_id,
            "full_name": row.full_name,
            "date_of_birth": dob_val.strftime("%Y-%m-%d") if dob_val else None,
            "ten_don_vi": ten_don_vi,
            "ten_phong_ban": ten_phong_ban,
            "shift_date": row.shift_date.strftime("%Y-%m-%d") if row.shift_date else None,
            "shift_code": row.shift_code
        })

    return data


# ==========================================
# 3. IMPORT / EXPORT EXCEL PHÂN CÔNG (MATRIX)
# ==========================================
@router.get("/api/assignments/export_template")
def export_assignment_template(
    start_date: str = None, 
    end_date: str = None, 
    search: str = Query(None),          # Hứng từ khóa tìm kiếm từ Frontend
    department_id: int = Query(None),   # Hứng bộ lọc phòng ban từ Frontend
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)  # Bắt buộc: Lấy thông tin user hiện tại
):
    current_username = current_user.get("username")
    current_role = current_user.get("role", "user")
    
    # 1. XỬ LÝ NGÀY THÁNG
    if not start_date or not end_date:
        today = date.today()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    else:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except Exception:
            today = date.today()
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
    
    date_list = []
    curr = start
    while curr <= end:
        date_list.append(curr)
        curr += timedelta(days=1)

    # 2. THIẾT LẬP FILE EXCEL
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill(start_color="4361EE", fill_type="solid")
    
    ws1 = wb.active
    ws1.title = "PhanCongCaTruc"
    
    base_headers = ["STT", "ID", "HỌ VÀ TÊN", "NGÀY SINH", "ĐƠN VỊ", "PHÒNG BAN"]
    date_headers = [d.strftime("%Y-%m-%d") for d in date_list]
    ws1.append(base_headers + date_headers)
    
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = blue_fill
    
    Dept = aliased(OrganizationUnit)
    ParentUnit = aliased(OrganizationUnit)
    
    # 3. TRUY VẤN CƠ BẢN (CHỈ LẤY NHÂN VIÊN ACTIVE)
    query = db.query(
        Employee,
        Dept.unit_name.label("dept_name"),
        ParentUnit.unit_name.label("parent_name")
    ).outerjoin(
        Dept, Employee.department_id == Dept.id
    ).outerjoin(
        ParentUnit, Dept.parent_id == ParentUnit.id
    ).filter(Employee.status == 'active')

    # ==========================================
    # 4. LOGIC PHÂN QUYỀN (RBAC)
    # ==========================================
    if current_role == "admin":
        pass # Admin xem tất cả
    elif current_role == "manager":
        manager_dept_id = db.query(Employee.department_id).filter(Employee.username == current_username).scalar()
        if manager_dept_id:
            query = query.filter(Employee.department_id == manager_dept_id)
        else:
            query = query.filter(Employee.username == current_username)
    else:
        # User thường chỉ xuất được lịch của chính mình
        query = query.filter(Employee.username == current_username)

    # ==========================================
    # 5. LỌC THEO TÌM KIẾM VÀ PHÒNG BAN (TỪ FRONTEND)
    # ==========================================
    if search:
        query = query.filter(
            or_(
                Employee.full_name.ilike(f"%{search}%"),
                Employee.username.ilike(f"%{search}%")
            )
        )
    if department_id:
        query = query.filter(Employee.department_id == department_id)

    # Chốt danh sách dòng
    rows = query.order_by(Employee.id.desc()).all()

    # 6. LẤY DỮ LIỆU CA TRỰC ĐÃ GÁN
    current_assigns = db.query(ShiftAssignment).filter(
        ShiftAssignment.shift_date >= start,
        ShiftAssignment.shift_date <= end
    ).all()
    
    assign_map = {} 
    for a in current_assigns:
        assign_map[(a.employee_id, a.shift_date.strftime("%Y-%m-%d"))] = a.shift_code

    # 7. ĐỔ DỮ LIỆU VÀO EXCEL
    for i, (e, dept_name, parent_name) in enumerate(rows):
        dob = e.dob or e.date_of_birth
        dob_str = dob.strftime("%Y-%m-%d") if dob else ""
        
        ten_don_vi = parent_name or dept_name or ""
        ten_phong_ban = dept_name if parent_name else ""
        
        row_data = [
            i + 1,
            e.id,
            e.full_name,
            dob_str,
            ten_don_vi,
            ten_phong_ban
        ]
        
        for d in date_list:
            d_str = d.strftime("%Y-%m-%d")
            row_data.append(assign_map.get((e.id, d_str), ""))
            
        ws1.append(row_data)

    # Chỉnh formating cột Sheet 1
    ws1.column_dimensions['A'].width = 8   # STT
    ws1.column_dimensions['B'].width = 12  # ID
    ws1.column_dimensions['C'].width = 35  # HỌ VÀ TÊN
    ws1.column_dimensions['D'].width = 15  # NGÀY SINH
    ws1.column_dimensions['E'].width = 45  # ĐƠN VỊ
    ws1.column_dimensions['F'].width = 45  # PHÒNG BAN

    for i in range(len(date_list)):
        col_letter = openpyxl.utils.get_column_letter(7 + i)
        ws1.column_dimensions[col_letter].width = 15

    # 8. TẠO SHEET 2 (DANH MỤC CA TRỰC)
    ws2 = wb.create_sheet(title="DanhMucCaTruc")
    ws2.append(["MÃ CA TRỰC", "TÊN CA TRỰC", "GIỜ BẮT ĐẦU", "GIỜ KẾT THÚC", "GHI CHÚ"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="10B981", fill_type="solid")
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 50
        
    shifts = db.query(ShiftCategory).all()
    for s in shifts:
        ws2.append([
            s.shift_code, s.shift_name, 
            s.start_time.strftime("%H:%M") if s.start_time else "",
            s.end_time.strftime("%H:%M") if s.end_time else "",
            s.notes or ""
        ])

    # 9. TRẢ VỀ FILE EXCEL QUA API
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Mau_Phan_Cong_{start.strftime('%d%m')}_{end.strftime('%d%m')}.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/api/assignments/import")
async def import_assignments(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents))
        
        if "PhanCongCaTruc" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="File Excel không đúng định dạng mẫu (Thiếu sheet PhanCongCaTruc).")
            
        ws = wb["PhanCongCaTruc"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"status": "success", "message": "File không có dữ liệu."}
        
        header = rows[0]
        try:
            id_idx = header.index("ID")
        except ValueError:
            raise HTTPException(status_code=400, detail="Mẫu Excel không hợp lệ (Không tìm thấy cột 'ID').")
        
        date_cols = []
        for i in range(6, len(header)):
            val = header[i]
            if val:
                try:
                    if isinstance(val, datetime):
                        date_cols.append((i, val.date()))
                    else:
                        d_obj = datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
                        date_cols.append((i, d_obj))
                except Exception:
                    continue

        valid_shift_codes = {s.shift_code for s in db.query(ShiftCategory).all()}

        count = 0
        for row in rows[1:]:
            emp_id = row[id_idx]
            if not emp_id: continue
            
            employee = db.query(Employee).filter(Employee.id == emp_id).first()
            if not employee: continue
            
            for col_idx, s_date in date_cols:
                val = row[col_idx]
                shift_code = str(val).strip() if val else None
                
                if shift_code and shift_code not in valid_shift_codes:
                    shift_code = None
                
                existing = db.query(ShiftAssignment).filter(
                    ShiftAssignment.employee_id == employee.id,
                    ShiftAssignment.shift_date == s_date
                ).first()
                
                if shift_code:
                    if existing:
                        existing.shift_code = shift_code
                    else:
                        db.add(ShiftAssignment(
                            employee_id=employee.id,
                            shift_code=shift_code,
                            shift_date=s_date
                        ))
                    count += 1
                elif existing:
                    db.delete(existing)
                    count += 1
            
        db.commit()
        return {"status": "success", "message": f"Đã cập nhật thành công {count} ô phân công!"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Lỗi đọc file Excel: {str(e)}")