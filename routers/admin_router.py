# Sửa dòng import này (Thêm Request)
from fastapi import APIRouter, Depends, HTTPException, Request

# Thêm 2 dòng này để hỗ trợ trả về giao diện HTML (nếu bạn để các hàm render HTML ở file này)
from fastapi.templating import Jinja2Templates
templates = Jinja2Templates(directory="templates")

from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date, time
from collections import defaultdict
import openpyxl, io
from openpyxl.styles import Font, PatternFill, Alignment
import os

from database import get_db
from models import Employee, Attendance, LeaveRequest, ShiftCategory
from schemas import EmployeeCreate, LeaveSubmit, ExplainRequest
from config import DB_PATH
from routers.auth_router import get_current_user

router = APIRouter()

# 2. Thêm Route trả về file dashboard.html (Ở mục 5. API ROUTES)
@router.get("/")
@router.get("/dashboard")
def read_dashboard(request: Request): 
    return templates.TemplateResponse("dashboard.html", {"request": request})

# 3. Thêm API tính toán thống kê (Ở mục 6. API DATABASE)
@router.get("/api/stats")
def get_dashboard_stats(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    role = current_user.get("role", "user")
    current_username = current_user.get("username")

    # Tính toán mốc thời gian của ngày hôm nay
    today_start = datetime.combine(date.today(), time.min)
    today_end = datetime.combine(date.today(), time.max)

    # 1. KHỞI TẠO CÁC QUERY CƠ BẢN
    emp_query = db.query(Employee)
    
    att_query = db.query(Attendance).filter(
        Attendance.check_in_time >= today_start,
        Attendance.check_in_time <= today_end
    )
    
    unique_checkin_query = db.query(Attendance.username).filter(
        Attendance.check_in_time >= today_start,
        Attendance.check_in_time <= today_end
    )

    # 2. ÁP DỤNG LOGIC PHÂN QUYỀN (RBAC)
    if role == "user":
        # User: Chỉ lấy thống kê của chính mình
        emp_query = emp_query.filter(Employee.username == current_username)
        att_query = att_query.filter(Attendance.username == current_username)
        unique_checkin_query = unique_checkin_query.filter(Attendance.username == current_username)

    elif role == "manager":
        # Manager: Lấy thống kê của các nhân viên trong cùng phòng ban
        manager = db.query(Employee).filter(Employee.username == current_username).first()
        if not manager or manager.department_id is None:
            # Nếu manager không thuộc phòng ban nào -> Trả về 0 để an toàn
            emp_query = emp_query.filter(False)
            att_query = att_query.filter(False)
            unique_checkin_query = unique_checkin_query.filter(False)
        else:
            # Lấy danh sách username của nhân viên cùng phòng ban
            dept_users = [
                u[0] for u in db.query(Employee.username).filter(Employee.department_id == manager.department_id).all()
            ]
            emp_query = emp_query.filter(Employee.username.in_(dept_users))
            att_query = att_query.filter(Attendance.username.in_(dept_users))
            unique_checkin_query = unique_checkin_query.filter(Attendance.username.in_(dept_users))
            
    # Admin: Không cần filter thêm gì, query sẽ lấy toàn bộ database

    # 3. THỰC THI QUERY VÀ LẤY KẾT QUẢ
    total_employees = emp_query.count()
    
    # Lưu ý: Danh mục ca (ShiftCategory) là dữ liệu dùng chung (Global) nên không cần phân quyền
    total_shifts = db.query(ShiftCategory).count() 

    today_attendances = att_query.count()
    unique_checkins_today = unique_checkin_query.distinct().count()

    return {
        "total_employees": total_employees,
        "total_shifts": total_shifts,
        "today_attendances": today_attendances,
        "unique_checkins_today": unique_checkins_today
    }



# --- (Thêm vào phần 5. GIAO DIỆN QUẢN TRỊ) ---
@router.get("/calendar")
def read_calendar(request: Request): 
    return templates.TemplateResponse("calendar.html", {"request": request})

@router.post("/api/leave")
def create_leave_request(req: LeaveSubmit, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.username == req.username).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")
    
    new_leave = LeaveRequest(
        username=req.username,
        full_name=emp.full_name,
        leave_date=req.leave_date,
        reason=req.reason,
        approver=req.approver
    )
    db.add(new_leave)
    db.commit()
    return {"status": "success", "message": "Đã gửi đơn đăng ký thành công"}

@router.get("/api/reports/export")
def export_report_excel(month: int, year: int, db: Session = Depends(get_db)):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # 1. LẤY DỮ LIỆU TỪ DATABASE
    attendances = db.query(Attendance).filter(
        Attendance.check_in_time >= start_date,
        Attendance.check_in_time < end_date
    ).order_by(Attendance.check_in_time.asc()).all()

    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.leave_date >= start_date.date(),
        LeaveRequest.leave_date < end_date.date()
    ).all()

    employees = db.query(Employee).all()
    emp_dict = {e.username: e.full_name for e in employees}

    # 2. XỬ LÝ LOGIC CHẤM CÔNG (Gom nhóm theo Nhân viên -> Ngày)
    # Cấu trúc: data[username][day] = {"in": time, "out": time, "late": 0, "early": 0}
    att_data = defaultdict(lambda: defaultdict(lambda: {"in": None, "out": None, "late": 0, "early": 0}))
    
    for a in attendances:
        user = a.username
        day = a.check_in_time.day
        t = a.check_in_time.time()
        time_str = t.strftime("%H:%M")

        if t.hour < 12: # Sáng: Chấm vào
            if not att_data[user][day]["in"] or time_str < att_data[user][day]["in"]:
                att_data[user][day]["in"] = time_str
                att_data[user][day]["late"] = a.late_minutes
        else: # Chiều: Chấm ra
            if not att_data[user][day]["out"] or time_str > att_data[user][day]["out"]:
                att_data[user][day]["out"] = time_str
                att_data[user][day]["early"] = a.early_minutes

    # 3. TẠO FILE EXCEL VỚI OPENPYXL
    wb = openpyxl.Workbook()
    
    # Định dạng style (In đậm, Nền xanh cho Header)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    # ==========================================
    # TAB 1: BẢNG CHẤM CÔNG & CÔNG THỰC TẾ
    # ==========================================
    ws_att = wb.active
    ws_att.title = "Bảng Chấm Công"
    
    headers_att = ["MÃ NV", "HỌ VÀ TÊN", "NGÀY", "GIỜ VÀO", "GIỜ RA", "ĐI MUỘN (Phút)", "VỀ SỚM (Phút)", "CÔNG THỰC TẾ (Giờ)", "TỔNG CÔNG THÁNG"]
    ws_att.append(headers_att)
    
    for col_num, cell in enumerate(ws_att[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Đổ dữ liệu chấm công
    for user, days in att_data.items():
        full_name = emp_dict.get(user, "Unknown")
        total_work_hours_month = 0.0
        
        # Mảng tạm để lưu các dòng của user này (nhằm ghi tổng công sau khi tính xong)
        user_rows = [] 
        
        for day in sorted(days.keys()):
            d_data = days[day]
            date_str = f"{day:02d}/{month:02d}/{year}"
            
            val_in = d_data["in"] or "--:--"
            val_out = d_data["out"] or "--:--"
            
            # Tính công thực tế ngày hôm đó
            work_hours_today = 0.0
            if d_data["in"] and d_data["out"]:
                t_in = datetime.strptime(d_data["in"], "%H:%M")
                t_out = datetime.strptime(d_data["out"], "%H:%M")
                diff_seconds = (t_out - t_in).total_seconds()
                
                # Trừ 1 tiếng nghỉ trưa (3600s)
                work_seconds = diff_seconds - 3600
                if work_seconds > 0:
                    work_hours_today = round(work_seconds / 3600, 2) # Làm tròn 2 chữ số thập phân
            
            total_work_hours_month += work_hours_today
            
            user_rows.append([
                user, full_name, date_str, 
                val_in, val_out, 
                d_data["late"], d_data["early"], 
                work_hours_today
            ])
            
        # Ghi các dòng của user này vào Excel, gắn thêm cột Tổng công vào dòng đầu tiên của họ
        for i, row in enumerate(user_rows):
            if i == 0:
                row.append(round(total_work_hours_month, 2)) # Cột tổng công chỉ hiện ở dòng đầu tiên của NV đó cho đỡ rối
            else:
                row.append("") # Các dòng sau để trống cột tổng công
            ws_att.append(row)

    # ==========================================
    # TAB 2: LỊCH NGHỈ PHÉP / CÔNG TÁC
    # ==========================================
    ws_leave = wb.create_sheet(title="Lịch Nghỉ Phép")
    headers_leave = ["MÃ NV", "HỌ VÀ TÊN", "NGÀY NGHỈ", "LÝ DO", "NGƯỜI DUYỆT", "TRẠNG THÁI"]
    ws_leave.append(headers_leave)
    
    for col_num, cell in enumerate(ws_leave[1], 1):
        cell.font = header_font
        cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid") # Nền cam cho tab nghỉ phép
        cell.alignment = align_center

    for l in leaves:
        ws_leave.append([
            l.username,
            l.full_name,
            l.leave_date.strftime("%d/%m/%Y"),
            l.reason,
            l.approver or "Chưa có",
            l.status
        ])

    # Tự động căn chỉnh độ rộng cột cho đẹp
    for ws in [ws_att, ws_leave]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # 4. TRẢ FILE EXCEL VỀ CHO TRÌNH DUYỆT TẢI XUỐNG
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=Bao_Cao_Cham_Cong_T{month}_{year}.xlsx"
    }
    # Chú ý: Đổi media_type sang định dạng chuẩn của Excel
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

@router.get("/api/payroll")
def calculate_payroll(month: int, year: int, db: Session = Depends(get_db)):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # Lấy danh sách nhân viên và chấm công
    employees = db.query(Employee).all()
    attendances = db.query(Attendance).filter(
        Attendance.check_in_time >= start_date,
        Attendance.check_in_time < end_date
    ).all()

    # Gom nhóm chấm công: user -> day -> in/out
    att_data = defaultdict(lambda: defaultdict(lambda: {"in": None, "out": None, "late": 0, "early": 0}))
    for a in attendances:
        user = a.username
        day = a.check_in_time.day
        t = a.check_in_time.time()
        time_str = t.strftime("%H:%M")

        if t.hour < 12: # Sáng
            if not att_data[user][day]["in"] or time_str < att_data[user][day]["in"]:
                att_data[user][day]["in"] = time_str
                att_data[user][day]["late"] = a.late_minutes
        else: # Chiều
            if not att_data[user][day]["out"] or time_str > att_data[user][day]["out"]:
                att_data[user][day]["out"] = time_str
                att_data[user][day]["early"] = a.early_minutes

    payroll_result = []

    for emp in employees:
        total_hours = 0.0
        total_late_min = 0
        total_early_min = 0
        user_att = att_data.get(emp.username, {})

        for day, d_data in user_att.items():
            total_late_min += d_data["late"]
            total_early_min += d_data["early"]
            
            # Tính công thực tế (Trừ 1 tiếng nghỉ trưa)
            if d_data["in"] and d_data["out"]:
                t_in = datetime.strptime(d_data["in"], "%H:%M")
                t_out = datetime.strptime(d_data["out"], "%H:%M")
                work_seconds = (t_out - t_in).total_seconds() - 3600
                if work_seconds > 0:
                    total_hours += round(work_seconds / 3600, 2)

        # Công thức tính lương
        # 1. Lương cơ bản = Tổng giờ * Lương 1 giờ
        gross_salary = total_hours * emp.hourly_rate
        
        # 2. Khấu trừ đi muộn/về sớm (Quy ra tiền: Số phút * Lương 1 phút)
        minute_rate = emp.hourly_rate / 60
        deductions = (total_late_min + total_early_min) * minute_rate

        # 3. Thực lãnh = Lương cơ bản + Phụ cấp - Khấu trừ
        net_salary = gross_salary + emp.allowance - deductions
        if net_salary < 0: net_salary = 0

        payroll_result.append({
            "username": emp.username,
            "full_name": emp.full_name,
            "department": emp.department,
            "hourly_rate": emp.hourly_rate,
            "allowance": emp.allowance,
            "total_hours": round(total_hours, 2),
            "total_late": total_late_min,
            "total_early": total_early_min,
            "gross_salary": round(gross_salary),
            "deductions": round(deductions),
            "net_salary": round(net_salary)
        })

    return payroll_result