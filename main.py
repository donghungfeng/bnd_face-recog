import cv2
import numpy as np
import base64
import os
import uvicorn
import pickle
import io
import csv
from datetime import datetime, date, time
from typing import Optional, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, HTTPException, BackgroundTasks, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from deepface import DeepFace

from calendar import monthrange # Thêm cái này để tính số ngày trong tháng
from collections import defaultdict

# --- SQLALCHEMY IMPORTS ---
from sqlalchemy import create_engine, Column, Integer, String, Date, Time, DateTime, Text
from sqlalchemy.orm import sessionmaker, declarative_base, Session

app = FastAPI(title="HRM AI Face Recognition")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# 1. CẤU HÌNH THƯ MỤC LƯU TRỮ & STATIC FILES
# ==========================================
DB_PATH = "./face_db"          # Chứa ảnh gốc + file cache AI
HISTORY_PATH = "./history_db"  # Chứa ảnh full-frame chấm công hàng ngày
TEMPLATE_PATH = "./templates"  # Chứa các file giao diện quản trị HTML

for path in [DB_PATH, HISTORY_PATH, TEMPLATE_PATH]:
    if not os.path.exists(path):
        os.makedirs(path)

# Cho phép trình duyệt truy cập trực tiếp vào thư mục chứa ảnh lịch sử
app.mount("/history_db", StaticFiles(directory="history_db"), name="history_db")

# Khởi tạo Jinja2 để load giao diện quản trị
templates = Jinja2Templates(directory="templates")


# ==========================================
# 2. CẤU HÌNH DATABASE SQLITE (ORM)
# ==========================================
SQLALCHEMY_DATABASE_URL = "sqlite:///./hrm.db"

engine = create_engine(
    SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False}
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# --- Định nghĩa các Bảng (Tables) ---
class Employee(Base):
    __tablename__ = "employees"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True) # VD: NV01
    full_name = Column(String, index=True)
    phone = Column(String, nullable=True)
    dob = Column(Date, nullable=True)
    email = Column(String, nullable=True)
    department = Column(String, nullable=True)
    status = Column(String, default="active")
    notes = Column(Text, nullable=True)

    hourly_rate = Column(Integer, default=25000) # Lương 1 giờ (VNĐ)
    allowance = Column(Integer, default=0)       # Phụ cấp tháng (VNĐ)

class Shift(Base):
    __tablename__ = "shifts"
    id = Column(Integer, primary_key=True, index=True)
    shift_code = Column(String, unique=True, index=True) # VD: CA_SANG
    start_time = Column(Time)
    end_time = Column(Time)

class Attendance(Base):
    __tablename__ = "attendance"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, index=True)
    full_name = Column(String)
    check_in_time = Column(DateTime, default=datetime.now)
    image_path = Column(String, nullable=True) 
    
    # BỔ SUNG CÁC CỘT MỚI:
    late_minutes = Column(Integer, default=0)       # Số phút đi muộn
    early_minutes = Column(Integer, default=0)      # Số phút về sớm
    explanation_status = Column(String, default="") # Trạng thái: "", "Đã gửi", "Đã duyệt"
    explanation_reason = Column(Text, nullable=True)# Lý do giải trình

class LeaveRequest(Base):
    __tablename__ = "leave_requests"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, index=True)
    full_name = Column(String)
    leave_date = Column(Date, index=True)
    reason = Column(String)
    approver = Column(String, nullable=True) # Người duyệt
    status = Column(String, default="Chờ duyệt") # Trạng thái đơn

# Tạo file hrm.db và các bảng nếu chưa tồn tại
Base.metadata.create_all(bind=engine)

# Hàm Dependency lấy DB Session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ==========================================
# 3. CẤU HÌNH AI MODEL & RAM CACHE
# ==========================================
MODEL_NAME = "Facenet512"
CACHE_FILE = os.path.join(DB_PATH, "embeddings_cache.pkl") 
known_face_embeddings = {}

def save_cache():
    with open(CACHE_FILE, "wb") as f:
        pickle.dump(known_face_embeddings, f)
    print(f"-> Đã lưu cache {len(known_face_embeddings)} nhân viên xuống ổ cứng.")

def load_embeddings():
    global known_face_embeddings
    if os.path.exists(CACHE_FILE):
        print("-> Đang nạp dữ liệu từ Cache (nhanh)...")
        with open(CACHE_FILE, "rb") as f:
            known_face_embeddings = pickle.load(f)
    else:
        print("-> Không tìm thấy Cache, sẽ tạo mới...")
        known_face_embeddings = {}

    changed = False
    for filename in os.listdir(DB_PATH):
        if filename.endswith(".jpg"):
            user_id = os.path.splitext(filename)[0]
            if user_id not in known_face_embeddings:
                print(f"-> Phát hiện ảnh mới [{user_id}], đang mã hóa...")
                img_path = os.path.join(DB_PATH, filename)
                try:
                    embedding = DeepFace.represent(img_path=img_path, model_name=MODEL_NAME, enforce_detection=False)[0]["embedding"]
                    known_face_embeddings[user_id] = np.array(embedding)
                    changed = True
                except Exception as e:
                    print(f"Lỗi khi xử lý {filename}: {e}")
    
    if changed:
        save_cache()
    print(f"-> SẴN SÀNG! Đã nạp {len(known_face_embeddings)} nhân viên vào RAM.")

# Khởi chạy nạp dữ liệu ngay khi chạy file
load_embeddings()


# ==========================================
# 4. PYDANTIC SCHEMAS & HELPERS
# ==========================================
class FaceRequest(BaseModel):
    user_id: str = None
    image_base64: str           # Ảnh crop để gửi vào DeepFace
    full_image_base64: str = None # Ảnh full để lưu làm bằng chứng

class UnregisterRequest(BaseModel):
    user_id: str

class ExplainRequest(BaseModel):
    id: int
    reason: str

class EmployeeCreate(BaseModel):
    username: str
    full_name: str
    phone: Optional[str] = None
    dob: Optional[date] = None
    email: Optional[str] = None
    department: Optional[str] = None
    status: Optional[str] = "active"
    notes: Optional[str] = None
    hourly_rate: int = 25000
    allowance: int = 0

class ShiftCreate(BaseModel):
    shift_code: str
    start_time: time
    end_time: time

class LeaveSubmit(BaseModel):
    username: str
    leave_date: date
    reason: str
    approver: str = ""

def decode_base64(data: str):
    if "," in data: data = data.split(",")[1]
    img_bytes = base64.b64decode(data)
    nparr = np.frombuffer(img_bytes, np.uint8)
    return cv2.imdecode(nparr, cv2.IMREAD_COLOR)

def cosine_similarity(v1, v2):
    return np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))

def background_logging(user_id: str, img_full: np.ndarray):
    """
    Hàm xử lý ngầm: Giữ tối đa 2 bản ghi/người/ngày.
    Lần 1: Giữ nguyên (Giờ vào)
    Lần 2 trở đi: Cập nhật liên tục thành giờ muộn nhất (Giờ ra)
    """
    db = SessionLocal() 
    try:
        now = datetime.now()
        
        # 1. Xác định mốc thời gian của ngày hôm nay
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

        # 2. Tìm TẤT CẢ bản ghi của nhân viên này trong ngày hôm nay (Xếp từ cũ đến mới)
        records_today = db.query(Attendance).filter(
            Attendance.username == user_id,
            Attendance.check_in_time >= today_start,
            Attendance.check_in_time <= today_end
        ).order_by(Attendance.check_in_time.asc()).all()

        # 3. Lưu ảnh mới ra ổ đĩa
        timestamp_str = now.strftime("%Y%m%d_%H%M%S")
        success_filename = f"{user_id}_{timestamp_str}.jpg"
        success_filepath = os.path.join(HISTORY_PATH, success_filename)
        cv2.imwrite(success_filepath, img_full)

        # 4. Tìm tên thật của nhân viên
        employee = db.query(Employee).filter(Employee.username == user_id).first()
        full_name = employee.full_name if employee else "Chưa cập nhật tên"

        # 5. Logic tính Đi Muộn / Về Sớm
        late_min = 0
        early_min = 0
        time_7_30 = now.replace(hour=7, minute=30, second=0, microsecond=0)
        time_17_00 = now.replace(hour=17, minute=0, second=0, microsecond=0)
        time_12_00 = now.replace(hour=12, minute=0, second=0, microsecond=0)

        if now < time_12_00:
            if now > time_7_30:
                late_min = int((now - time_7_30).total_seconds() / 60)
        else:
            if now < time_17_00:
                early_min = int((time_17_00 - now).total_seconds() / 60)

        image_web_path = f"/history_db/{success_filename}"

        # =======================================================
        # 6. LOGIC GIỮ TỐI ĐA 2 BẢN GHI
        # =======================================================
        if len(records_today) == 0:
            # TRƯỜNG HỢP 1: Quẹt lần đầu tiên trong ngày -> TẠO MỚI (Lần 1)
            new_log = Attendance(
                username=user_id,
                full_name=full_name,
                check_in_time=now,
                image_path=image_web_path,
                late_minutes=late_min,
                early_minutes=early_min
            )
            db.add(new_log)

        elif len(records_today) == 1:
            # TRƯỜNG HỢP 2: Quẹt lần thứ hai trong ngày -> TẠO MỚI (Lần 2)
            new_log = Attendance(
                username=user_id,
                full_name=full_name,
                check_in_time=now,
                image_path=image_web_path,
                late_minutes=late_min,
                early_minutes=early_min
            )
            db.add(new_log)

        else:
            # TRƯỜNG HỢP 3: Quẹt từ lần thứ 3 trở lên -> GHI ĐÈ LÊN BẢN GHI THỨ 2
            # Bản ghi thứ 1 (records_today[0]) LUÔN ĐƯỢC GIỮ NGUYÊN (Sớm nhất)
            latest_record = records_today[1] 

            # Xóa file ảnh cũ của bản ghi thứ 2 khỏi ổ cứng để tiết kiệm dung lượng
            if latest_record.image_path:
                old_img_path = "." + latest_record.image_path 
                if os.path.exists(old_img_path):
                    os.remove(old_img_path)

            # Cập nhật thông tin giờ & ảnh mới nhất
            latest_record.check_in_time = now
            latest_record.image_path = image_web_path
            latest_record.late_minutes = late_min
            latest_record.early_minutes = early_min

            # Đề phòng rủi ro database cũ bị lỗi có nhiều hơn 2 bản ghi, dọn dẹp sạch sẽ
            if len(records_today) > 2:
                for extra_record in records_today[2:]:
                    if extra_record.image_path:
                        extra_img_path = "." + extra_record.image_path
                        if os.path.exists(extra_img_path):
                            os.remove(extra_img_path)
                    db.delete(extra_record)

        # Lưu thay đổi vào DB
        db.commit()

    except Exception as e:
        print(f"Lỗi ghi log ngầm: {e}")
    finally:
        db.close()

# ==========================================
# 5. API ROUTES (GIAO DIỆN & AI)
# ==========================================

# --- GIAO DIỆN IPAD CHẤM CÔNG VÀ ĐĂNG KÝ (Trả về file trực tiếp) ---
@app.get("/verify")
def read_index(): 
    return FileResponse("verify.html")

@app.get("/enroll")
def read_enroll(): 
    return FileResponse("enroll.html")

# --- GIAO DIỆN QUẢN TRỊ ADMIN (Render qua Jinja2 Templates) ---
@app.get("/employees")
def read_employees(request: Request): 
    return templates.TemplateResponse("employees.html", {"request": request})

@app.get("/shifts")
def read_shifts(request: Request): 
    return templates.TemplateResponse("shifts.html", {"request": request})

@app.get("/attendance")
def read_attendance(request: Request): 
    return templates.TemplateResponse("attendance.html", {"request": request})

@app.get("/payroll")
def read_payroll(request: Request): 
    return templates.TemplateResponse("payroll.html", {"request": request})


# --- CÁC API NHẬN DIỆN KHUÔN MẶT ---
@app.post("/register")
async def register(request: FaceRequest):
    img = decode_base64(request.image_base64)
    file_path = os.path.join(DB_PATH, f"{request.user_id}.jpg")
    cv2.imwrite(file_path, img)
    
    try:
        # Cập nhật ngay RAM và Cache mà không cần load lại toàn bộ thư mục
        embedding = DeepFace.represent(img_path=img, model_name=MODEL_NAME, enforce_detection=False)[0]["embedding"]
        known_face_embeddings[request.user_id] = np.array(embedding)
        save_cache() 
    except Exception as e:
        return {"status": "error", "message": f"Lỗi AI khi mã hóa: {e}"}
        
    return {"status": "success", "message": f"Đã đăng ký {request.user_id}"}


@app.post("/recognize")
async def recognize(request: FaceRequest, background_tasks: BackgroundTasks):
    img_crop = decode_base64(request.image_base64)
    
    # Ưu tiên lấy ảnh full nếu Client có gửi
    if request.full_image_base64:
        img_to_save = decode_base64(request.full_image_base64)
    else:
        img_to_save = img_crop

    try:
        results = DeepFace.represent(img_path=img_crop, model_name=MODEL_NAME, enforce_detection=False)
        if not results:
            return {"recognized": False, "message": "Không thấy mặt"}
        
        current_embedding = np.array(results[0]["embedding"])
        best_match = None
        max_sim = -1

        for user_id, known_emb in known_face_embeddings.items():
            sim = cosine_similarity(current_embedding, known_emb)
            if sim > max_sim:
                max_sim = sim
                best_match = user_id

        THRESHOLD = 0.75

        if max_sim >= THRESHOLD:
            # GỌI HÀM CHẠY NGẦM: Không bắt API chờ việc ghi đĩa
            background_tasks.add_task(background_logging, best_match, img_to_save)
            
            return {"recognized": True, "user_id": best_match, "match_probability": f"{round(max_sim * 100, 2)}%"}
        else:
            return {"recognized": False, "message": "Người lạ", "match_probability": f"{round(max_sim * 100, 2)}%"}

    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/unregister")
async def unregister(request: UnregisterRequest):
    file_path = os.path.join(DB_PATH, f"{request.user_id}.jpg")
    if os.path.exists(file_path):
        try:
            os.remove(file_path) 
            pkl_path = os.path.join(DB_PATH, f"representations_{MODEL_NAME.lower()}.pkl")
            if os.path.exists(pkl_path): os.remove(pkl_path)
                
            if request.user_id in known_face_embeddings:
                del known_face_embeddings[request.user_id]
                save_cache() 
            return {"status": "success", "message": f"Đã hủy đăng ký: {request.user_id}"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Lỗi khi xóa dữ liệu: {str(e)}")
    else:
        return {"status": "error", "message": "Không tìm thấy người này."}


@app.get("/clear_ram")
async def clear_ram():
    global known_face_embeddings
    known_face_embeddings = {}
    if os.path.exists(CACHE_FILE):
        os.remove(CACHE_FILE)
    return {"status": "success", "message": "Hệ thống đã được dọn sạch RAM & Cache."}


# ==========================================
# 6. API DATABASE (CRUD - Dùng cho các trang quản trị)
# ==========================================

@app.post("/api/employees")
def create_employee(emp: EmployeeCreate, db: Session = Depends(get_db)):
    db_emp = db.query(Employee).filter(Employee.username == emp.username).first()
    if db_emp:
        raise HTTPException(status_code=400, detail="Mã nhân sự đã tồn tại")
    
    new_emp = Employee(**emp.dict())
    db.add(new_emp)
    db.commit()
    db.refresh(new_emp)
    return {"status": "success", "message": "Thêm nhân sự thành công"}

@app.get("/api/employees")
def get_employees(db: Session = Depends(get_db)):
    emps = db.query(Employee).all()
    result = []
    for e in emps:
        # Kiểm tra file ảnh khuôn mặt đã tồn tại chưa
        face_path = os.path.join(DB_PATH, f"{e.username}.jpg")
        has_face = os.path.exists(face_path)
        
        result.append({
            "username": e.username,
            "full_name": e.full_name,
            "department": e.department,
            "phone": e.phone,
            "dob": e.dob,
            "status": e.status,
            "has_face": has_face
        })
    return result

@app.put("/api/employees/{username}")
def update_employee(username: str, emp: EmployeeCreate, db: Session = Depends(get_db)):
    db_emp = db.query(Employee).filter(Employee.username == username).first()
    if not db_emp:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân sự")
    
    db_emp.username = emp.username
    db_emp.full_name = emp.full_name
    db_emp.phone = emp.phone
    db_emp.dob = emp.dob
    db_emp.email = emp.email
    db_emp.department = emp.department
    db_emp.status = emp.status
    db_emp.notes = emp.notes

    db.commit()
    return {"status": "success", "message": "Đã cập nhật thông tin nhân sự"}

@app.delete("/api/employees/{username}")
def delete_employee(username: str, db: Session = Depends(get_db)):
    db_emp = db.query(Employee).filter(Employee.username == username).first()
    if not db_emp:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân sự")
    
    db.delete(db_emp)
    db.commit()
    return {"status": "success", "message": "Đã xóa nhân sự thành công"}


@app.post("/api/shifts")
def create_shift(shift: ShiftCreate, db: Session = Depends(get_db)):
    db_shift = db.query(Shift).filter(Shift.shift_code == shift.shift_code).first()
    if db_shift:
        raise HTTPException(status_code=400, detail="Mã ca làm việc đã tồn tại")
    
    new_shift = Shift(**shift.dict())
    db.add(new_shift)
    db.commit()
    db.refresh(new_shift)
    return {"status": "success", "message": "Thêm ca làm việc thành công"}

@app.get("/api/shifts")
def get_shifts(db: Session = Depends(get_db)):
    return db.query(Shift).all()

@app.get("/api/attendance")
def get_attendance(limit: int = 100, db: Session = Depends(get_db)):
    # Trả về tối đa 'limit' dòng mới nhất
    return db.query(Attendance).order_by(Attendance.check_in_time.desc()).limit(limit).all()

# 2. Thêm Route trả về file dashboard.html (Ở mục 5. API ROUTES)
@app.get("/")
@app.get("/dashboard")
def read_dashboard(request: Request): 
    return templates.TemplateResponse("dashboard.html", {"request": request})

# 3. Thêm API tính toán thống kê (Ở mục 6. API DATABASE)
@app.get("/api/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    # Đếm tổng nhân sự & tổng ca
    total_employees = db.query(Employee).count()
    total_shifts = db.query(Shift).count()

    # Tính toán mốc thời gian của ngày hôm nay
    today_start = datetime.combine(date.today(), time.min)
    today_end = datetime.combine(date.today(), time.max)

    # Đếm tổng số lượt chấm công hôm nay
    today_attendances = db.query(Attendance).filter(
        Attendance.check_in_time >= today_start,
        Attendance.check_in_time <= today_end
    ).count()

    # Đếm số lượng nhân viên thực tế đã đi làm hôm nay (Loại bỏ trùng lặp nếu 1 người quẹt nhiều lần)
    unique_checkins_today = db.query(Attendance.username).filter(
        Attendance.check_in_time >= today_start,
        Attendance.check_in_time <= today_end
    ).distinct().count()

    return {
        "total_employees": total_employees,
        "total_shifts": total_shifts,
        "today_attendances": today_attendances,
        "unique_checkins_today": unique_checkins_today
    }

@app.post("/api/attendance/explain")
def submit_explanation(req: ExplainRequest, db: Session = Depends(get_db)):
    record = db.query(Attendance).filter(Attendance.id == req.id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")
    
    record.explanation_reason = req.reason
    record.explanation_status = "Đã gửi" # Đổi trạng thái
    db.commit()
    return {"status": "success", "message": "Đã gửi giải trình thành công"}

# --- (Thêm vào phần 5. GIAO DIỆN QUẢN TRỊ) ---
@app.get("/calendar")
def read_calendar(request: Request): 
    return templates.TemplateResponse("calendar.html", {"request": request})


# --- (Thêm vào phần 6. API DATABASE) ---
@app.get("/api/attendance/calendar")
def get_calendar_data(username: str, month: int, year: int, db: Session = Depends(get_db)):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # 1. Lấy dữ liệu chấm công
    records = db.query(Attendance).filter(
        Attendance.username == username,
        Attendance.check_in_time >= start_date,
        Attendance.check_in_time < end_date
    ).all()

    # 2. Lấy dữ liệu xin nghỉ phép/công tác
    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.username == username,
        LeaveRequest.leave_date >= start_date.date(),
        LeaveRequest.leave_date < end_date.date()
    ).all()

    daily_data = defaultdict(lambda: {"in": None, "out": None, "work_time": None, "leave": None})

    # Đưa dữ liệu nghỉ phép vào ngày tương ứng
    for l in leaves:
        daily_data[l.leave_date.day]["leave"] = {
            "reason": l.reason,
            "status": l.status,
            "approver": l.approver
        }

    # Đưa dữ liệu Giờ Vào/Ra vào ngày
    for r in records:
        day = r.check_in_time.day
        t = r.check_in_time.time()
        time_str = t.strftime("%H:%M")

        if t.hour < 12:
            if not daily_data[day]["in"] or time_str < daily_data[day]["in"]:
                daily_data[day]["in"] = time_str
        else:
            if not daily_data[day]["out"] or time_str > daily_data[day]["out"]:
                daily_data[day]["out"] = time_str

    # 3. Tính toán CÔNG THỰC TẾ = Ra - Vào - 1 tiếng (nghỉ trưa)
    for day, data in daily_data.items():
        if data["in"] and data["out"]:
            t_in = datetime.strptime(data["in"], "%H:%M")
            t_out = datetime.strptime(data["out"], "%H:%M")
            diff_seconds = (t_out - t_in).total_seconds()
            
            work_seconds = diff_seconds - 3600 # Trừ đi 3600 giây (1 tiếng nghỉ trưa)
            
            if work_seconds > 0:
                h = int(work_seconds // 3600)
                m = int((work_seconds % 3600) // 60)
                data["work_time"] = f"{h}h{m}p"
            else:
                data["work_time"] = "0h"

    return daily_data

@app.post("/api/leave")
def create_leave_request(req: LeaveSubmit, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.username == req.username).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhân viên")
    
    new_leave = LeaveRequest(
        username=req.username,
        full_name=emp.full_name,
        leave_date=req.leave_date,
        reason=req.reason,
        approver=req.approver
    )
    db.add(new_leave)
    db.commit()
    return {"status": "success", "message": "Đã gửi đơn đăng ký thành công"}

@app.get("/api/reports/export")
def export_report_excel(month: int, year: int, db: Session = Depends(get_db)):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # 1. LẤY DỮ LIỆU TỪ DATABASE
    attendances = db.query(Attendance).filter(
        Attendance.check_in_time >= start_date,
        Attendance.check_in_time < end_date
    ).order_by(Attendance.check_in_time.asc()).all()

    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.leave_date >= start_date.date(),
        LeaveRequest.leave_date < end_date.date()
    ).all()

    employees = db.query(Employee).all()
    emp_dict = {e.username: e.full_name for e in employees}

    # 2. XỬ LÝ LOGIC CHẤM CÔNG (Gom nhóm theo Nhân viên -> Ngày)
    # Cấu trúc: data[username][day] = {"in": time, "out": time, "late": 0, "early": 0}
    att_data = defaultdict(lambda: defaultdict(lambda: {"in": None, "out": None, "late": 0, "early": 0}))
    
    for a in attendances:
        user = a.username
        day = a.check_in_time.day
        t = a.check_in_time.time()
        time_str = t.strftime("%H:%M")

        if t.hour < 12: # Sáng: Chấm vào
            if not att_data[user][day]["in"] or time_str < att_data[user][day]["in"]:
                att_data[user][day]["in"] = time_str
                att_data[user][day]["late"] = a.late_minutes
        else: # Chiều: Chấm ra
            if not att_data[user][day]["out"] or time_str > att_data[user][day]["out"]:
                att_data[user][day]["out"] = time_str
                att_data[user][day]["early"] = a.early_minutes

    # 3. TẠO FILE EXCEL VỚI OPENPYXL
    wb = openpyxl.Workbook()
    
    # Định dạng style (In đậm, Nền xanh cho Header)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    # ==========================================
    # TAB 1: BẢNG CHẤM CÔNG & CÔNG THỰC TẾ
    # ==========================================
    ws_att = wb.active
    ws_att.title = "Bảng Chấm Công"
    
    headers_att = ["MÃ NV", "HỌ VÀ TÊN", "NGÀY", "GIỜ VÀO", "GIỜ RA", "ĐI MUỘN (Phút)", "VỀ SỚM (Phút)", "CÔNG THỰC TẾ (Giờ)", "TỔNG CÔNG THÁNG"]
    ws_att.append(headers_att)
    
    for col_num, cell in enumerate(ws_att[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Đổ dữ liệu chấm công
    for user, days in att_data.items():
        full_name = emp_dict.get(user, "Unknown")
        total_work_hours_month = 0.0
        
        # Mảng tạm để lưu các dòng của user này (nhằm ghi tổng công sau khi tính xong)
        user_rows = [] 
        
        for day in sorted(days.keys()):
            d_data = days[day]
            date_str = f"{day:02d}/{month:02d}/{year}"
            
            val_in = d_data["in"] or "--:--"
            val_out = d_data["out"] or "--:--"
            
            # Tính công thực tế ngày hôm đó
            work_hours_today = 0.0
            if d_data["in"] and d_data["out"]:
                t_in = datetime.strptime(d_data["in"], "%H:%M")
                t_out = datetime.strptime(d_data["out"], "%H:%M")
                diff_seconds = (t_out - t_in).total_seconds()
                
                # Trừ 1 tiếng nghỉ trưa (3600s)
                work_seconds = diff_seconds - 3600
                if work_seconds > 0:
                    work_hours_today = round(work_seconds / 3600, 2) # Làm tròn 2 chữ số thập phân
            
            total_work_hours_month += work_hours_today
            
            user_rows.append([
                user, full_name, date_str, 
                val_in, val_out, 
                d_data["late"], d_data["early"], 
                work_hours_today
            ])
            
        # Ghi các dòng của user này vào Excel, gắn thêm cột Tổng công vào dòng đầu tiên của họ
        for i, row in enumerate(user_rows):
            if i == 0:
                row.append(round(total_work_hours_month, 2)) # Cột tổng công chỉ hiện ở dòng đầu tiên của NV đó cho đỡ rối
            else:
                row.append("") # Các dòng sau để trống cột tổng công
            ws_att.append(row)

    # ==========================================
    # TAB 2: LỊCH NGHỈ PHÉP / CÔNG TÁC
    # ==========================================
    ws_leave = wb.create_sheet(title="Lịch Nghỉ Phép")
    headers_leave = ["MÃ NV", "HỌ VÀ TÊN", "NGÀY NGHỈ", "LÝ DO", "NGƯỜI DUYỆT", "TRẠNG THÁI"]
    ws_leave.append(headers_leave)
    
    for col_num, cell in enumerate(ws_leave[1], 1):
        cell.font = header_font
        cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid") # Nền cam cho tab nghỉ phép
        cell.alignment = align_center

    for l in leaves:
        ws_leave.append([
            l.username,
            l.full_name,
            l.leave_date.strftime("%d/%m/%Y"),
            l.reason,
            l.approver or "Chưa có",
            l.status
        ])

    # Tự động căn chỉnh độ rộng cột cho đẹp
    for ws in [ws_att, ws_leave]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # 4. TRẢ FILE EXCEL VỀ CHO TRÌNH DUYỆT TẢI XUỐNG
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=Bao_Cao_Cham_Cong_T{month}_{year}.xlsx"
    }
    # Chú ý: Đổi media_type sang định dạng chuẩn của Excel
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

@app.get("/api/payroll")
def calculate_payroll(month: int, year: int, db: Session = Depends(get_db)):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # Lấy danh sách nhân viên và chấm công
    employees = db.query(Employee).all()
    attendances = db.query(Attendance).filter(
        Attendance.check_in_time >= start_date,
        Attendance.check_in_time < end_date
    ).all()

    # Gom nhóm chấm công: user -> day -> in/out
    att_data = defaultdict(lambda: defaultdict(lambda: {"in": None, "out": None, "late": 0, "early": 0}))
    for a in attendances:
        user = a.username
        day = a.check_in_time.day
        t = a.check_in_time.time()
        time_str = t.strftime("%H:%M")

        if t.hour < 12: # Sáng
            if not att_data[user][day]["in"] or time_str < att_data[user][day]["in"]:
                att_data[user][day]["in"] = time_str
                att_data[user][day]["late"] = a.late_minutes
        else: # Chiều
            if not att_data[user][day]["out"] or time_str > att_data[user][day]["out"]:
                att_data[user][day]["out"] = time_str
                att_data[user][day]["early"] = a.early_minutes

    payroll_result = []

    for emp in employees:
        total_hours = 0.0
        total_late_min = 0
        total_early_min = 0
        user_att = att_data.get(emp.username, {})

        for day, d_data in user_att.items():
            total_late_min += d_data["late"]
            total_early_min += d_data["early"]
            
            # Tính công thực tế (Trừ 1 tiếng nghỉ trưa)
            if d_data["in"] and d_data["out"]:
                t_in = datetime.strptime(d_data["in"], "%H:%M")
                t_out = datetime.strptime(d_data["out"], "%H:%M")
                work_seconds = (t_out - t_in).total_seconds() - 3600
                if work_seconds > 0:
                    total_hours += round(work_seconds / 3600, 2)

        # Công thức tính lương
        # 1. Lương cơ bản = Tổng giờ * Lương 1 giờ
        gross_salary = total_hours * emp.hourly_rate
        
        # 2. Khấu trừ đi muộn/về sớm (Quy ra tiền: Số phút * Lương 1 phút)
        minute_rate = emp.hourly_rate / 60
        deductions = (total_late_min + total_early_min) * minute_rate

        # 3. Thực lãnh = Lương cơ bản + Phụ cấp - Khấu trừ
        net_salary = gross_salary + emp.allowance - deductions
        if net_salary < 0: net_salary = 0

        payroll_result.append({
            "username": emp.username,
            "full_name": emp.full_name,
            "department": emp.department,
            "hourly_rate": emp.hourly_rate,
            "allowance": emp.allowance,
            "total_hours": round(total_hours, 2),
            "total_late": total_late_min,
            "total_early": total_early_min,
            "gross_salary": round(gross_salary),
            "deductions": round(deductions),
            "net_salary": round(net_salary)
        })

    return payroll_result

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)